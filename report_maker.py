import pandas as pd
import numpy as np
import openpyxl
import os
from openpyxl.chart import LineChart, Reference, Series

from datetime import datetime, timedelta, time
from tqdm import tqdm
from pandas_ods_reader import read_ods

months = ["1 month", "3 month", "12 month"]

def determineTrade(previous, current, threshold):

  if previous * (100 + threshold) / 100 - current < 0:
    return "buy"

  elif previous * (100 - threshold) / 100 - current > 0:
    return "sell"

  else:
    return "keep"

def insertValue(out_dict, buy_count, sell_count, keep_count):
  out_dict["Buy Count"].append(buy_count[0])
  out_dict["Sell Count"].append(sell_count[0])
  out_dict["No trade"].append(keep_count[0])
  out_dict["Buy PNL"].append(buy_count[1])
  out_dict["Sell PNL"].append(sell_count[1])
  out_dict["Total PNL"].append(keep_count[1])
  out_dict["Average"].append(keep_count[2])

def returnDict():
  return {
    "Buy Count": [],
    "Sell Count": [],
    "No trade": [],
    "Buy PNL": [],
    "Sell PNL": [],
    "Total PNL": [],
    "Average": []
  }

def calculateValues(check_points, out_dict, data, single_point, settlement, pre_settlement, threshold):


  buy_count = [0, 0, 0]
  sell_count = [0, 0, 0]
  keep_count = [0, 0, 0]

  day_before_price = None
  hour_before_price = None

  trade = "keep"
  _ = 0

  for index in tqdm(range(len(data)), desc="processing..."):

    row = data.iloc[index, :]
    check_point = row.tolist()
    check_point.append("")
    check_point.append("")
    check_point.append("")
    check_point.append("")

    time_str = row["timestamp"]

    if time_str[-8:] == pre_settlement:
      if day_before_price != None:
        hour_before_price = float(row["close"])
        trade = determineTrade(day_before_price, hour_before_price, threshold)
        
        check_point[4] = trade

        if trade != "keep":
          _ += 1
          check_point[3] = int(single_point)
          check_point[6] = float(buy_count[1] + sell_count[1])

          check_points.append(check_point)
  
    elif time_str[-8:] == settlement:
      check_point[4] = trade

      if time_str[:10] == str(start_date + timedelta(30))[:10]:
        keep_count[1] = buy_count[1] + sell_count[1]
        _counts = buy_count[0] + sell_count[0]
        _value = keep_count[1] / _counts if _counts != 0 else 0

        keep_count[2] = round(_value, 3)
        insertValue(out_dict, buy_count, sell_count, keep_count)

      elif time_str[:10] == str(start_date + timedelta(90))[:10]:
        keep_count[1] = buy_count[1] + sell_count[1]
        _counts = buy_count[0] + sell_count[0]
        _value = keep_count[1] / _counts if _counts != 0 else 0
        
        keep_count[2] = round(_value, 3)
        insertValue(out_dict, buy_count, sell_count, keep_count)

      if trade == "buy":
        buy_count[0] += 1
        buy_count[1] += int((float(row["close"]) - hour_before_price) * single_point)

        check_point[4] = "result"
        check_point[5] = int((float(row["close"]) - hour_before_price) * single_point)

      elif trade == "sell":
        sell_count[0] += 1
        sell_count[1] += int((hour_before_price - float(row["close"])) * single_point)

        check_point[4] = "result"
        check_point[5] = int((hour_before_price - float(row["close"])) * single_point)

      else:
        keep_count[0] += 1

      trade = "keep"
      day_before_price = float(row["close"])

      check_point[3] = int(single_point)
      check_point[6] = float(buy_count[1] + sell_count[1])

      check_points.append(check_point)
      _ += 1

  keep_count[1] = buy_count[1] + sell_count[1]
  _counts = buy_count[0] + sell_count[0]
  _value = keep_count[1] / _counts if _counts != 0 else 0
        
  keep_count[2] = round(_value, 3)

  insertValue(out_dict, buy_count, sell_count, keep_count)

  return _

def writeResult(_dict):
  ## wirte the result as xlsx file using openpyxl.

  print("writing...")
  def _get_column_num(number):

    new_number = number
    new_number += 4 if number > 0 else 0
    new_number += 1 if number > 4 else 0
    new_number += 1 if number > 7 else 0
    return new_number

  book = openpyxl.Workbook()
  sheet = book.active

  symbol_count = len(_dict.keys())

  ## write the first row.
  for index, key in enumerate(returnDict().keys()):
    sheet.cell(row=2, column=_get_column_num(index + 1 + 1)).value = key

  ## write each values.
  for symbol_index, symbol_key in enumerate(_dict):
    for column_index, column_key in enumerate(_dict[symbol_key]):
      basic_row = symbol_index + 2 + 1
      basic_col = _get_column_num(column_index + 1 + 1)

      for index, item in enumerate(_dict[symbol_key][column_key]):
        real_row = basic_row + index * (symbol_count + 2)
        sheet.cell(row=real_row, column=basic_col).value = item

        if column_index == 0:
          sheet.cell(row=real_row, column=_get_column_num(1)).value = symbol_key

          if symbol_index == 0:
            sheet.cell(row=real_row - 1, column=_get_column_num(1)).value = months[index]

  ## write to file.
  result_path = out_path + "/result.xlsx"

  if os.path.exists(result_path):
    os.remove(result_path)

  os.makedirs(out_path, exist_ok=True)
  
  book.save(result_path)
      

if __name__ == "__main__":
  ## path for the input file.
  root_path = "./input"
  out_path = "./output"
  path = "/AutoData/iqfeed_hist_data_20200205-20210205.csv"

  ## prepare the source raw data.
  df_data = pd.read_csv(root_path + path)
  df_data = df_data[["timestamp", "close", "symbol"]]

  ## prepare BuyStrength_SellWeakness data.
  df_additional = read_ods(root_path + "/BuyStrength_SellWeakness.ods", 1)
  df_additional = df_additional.iloc[2:, :]
  df_additional = df_additional[df_additional["unnamed.2"].notnull()].reset_index(drop=True)

  ## prepare IQFeed_Symbols data.
  df_singlepoint = pd.read_excel(root_path + "/IQFeed_Symbols.xlsx", engine="openpyxl")
  df_singlepoint = df_singlepoint[df_singlepoint["Symbol"].notnull()].reset_index(drop=True)

  ## prepare variables to save data.
  check_points = []
  precious_dict = {}
  item_lens = []

  ## set start date.
  start_date = datetime(2020, 2, 5)

  for index in tqdm(range(len(df_additional.index))):

    row = df_additional.iloc[index, :]
    _time_str = row.iloc[4]
    _product_name = row.iloc[1]
    _row_single = df_singlepoint[df_singlepoint["Name"] == _product_name]
    _product_code = row.iloc[0]

    _hour = int(_time_str[2:_time_str.find('H')])
    _min = int(_time_str[_time_str.find('H') + 1: _time_str.find('M')])

    settlement = str(time(_hour, _min))
    pre_settlement = str(time(23 if _hour == 0 else _hour - 1, _min))
    symbol = _row_single["Symbol"].values[0]
    single_point = _row_single["Single point value"].values[0]
    threshold = row.iloc[5]
    threshold = threshold * 100

    precious_dict.update({_product_code: returnDict()})
    data = df_data[df_data["symbol"] == symbol].reset_index(drop=True)

    item_lens.append([_product_code, calculateValues(check_points, precious_dict[_product_code], data, single_point, settlement, pre_settlement, threshold)])

  writeResult(precious_dict)

  ## output checkpoint file.
  df = pd.DataFrame(check_points, columns=["timestamp", "close", "symbol", "single point", "event", "result", "profit"])
  df.to_excel(out_path + "/check_points.xlsx", index=False)

  ## draw graph on checkpoint file.
  book = openpyxl.load_workbook(out_path + "/check_points.xlsx")
  sheet = book.active

  row_over = 0
  for index, item in enumerate(item_lens):

    chart = LineChart()
    chart.title = item[0]
    chart.x_axis.title = "Time"
    chart.y_axis.title = "profit"
    chart.style = 2
    values = Reference(sheet, min_col=6, min_row=row_over+2, max_col=7, max_row=row_over+item[1]+1)
    chart.add_data(values)
    chart.height = 10

    sheet.add_chart(chart, "I" + str(5+row_over))
    row_over += item[1]

  book.save(out_path + "/check_points.xlsx")

  print("done!")