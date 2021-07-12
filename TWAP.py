import pandas as pd
import numpy as np
import openpyxl
import os
from openpyxl.chart import LineChart, Reference, Series

from datetime import datetime, timedelta, time
from tqdm import tqdm
from pandas_ods_reader import read_ods

durations = [5, 21, 63, 242]
months = ["last {} days".format(durations[0]), 
            "last {} days".format(durations[1]), 
            "last {} days".format(durations[2]), 
            "last {} days".format(durations[3]), 
            "last year"]

def insertValue(out_dict, store):
  out_dict["TWAP"].append(store)

def returnDict():
  return {
    "TWAP": [],
  }

def calculateValues(weights, check_points, graph_points, out_dict, data, single_point, settlement, open_time, threshold, symbol, summary_stat):

  reduce_day = []
  days = 0
  reduce_interval = 0
  count_interval = 0

  delta_list = []

  stack = []

  _time = settlement
  _settle_hour = int(_time[:_time.index(':')])
  _settle_min = int(_time[_time.index(':') + 1:_time.index(':', 3)])   
  settle_time = time(_settle_hour, _settle_min)

  for index in tqdm(range(len(data)), desc="processing..."):

    row = data.iloc[index, :]
    graph_point = row.tolist()
    check_point = row.tolist()

    graph_point.append("")
    graph_point.append("")

    check_point.append("")
    check_point.append("")
    check_point.append("")
    check_point.append("")
    check_point.append("")
    check_point.append("")
    
    time_str = row["timestamp"]
    _day_str = time_str[:-9]
    _year = int(_day_str[: _day_str.index('-')])
    _month = int(_day_str[_day_str.index('-') + 1: _day_str.index('-', 5)])
    _day = int(_day_str[_day_str.index('-', 5) + 1:])

    _time = time_str[-8:]
    _hour = int(_time[:_time.index(':')])
    _min = int(_time[_time.index(':') + 1:_time.index(':', 3)])

    now_datetime = datetime(_year, _month, _day, _hour, _min)

    now_time = time(_hour, _min)
  
    if now_time > open_time or now_time < settle_time:

      value = (float(row["high"]) + float(row["low"]) + float(row["open"]) + float(row["close"])) / 4
      check_point[6] = value

      stack.append([now_datetime, value])

      reduce_interval += value
      count_interval += 1

    if time_str[-8:] == settlement:
      
      if count_interval != 0:

        days += 1

        _ = 0
        __ = 0
        for item in stack:
          if item[0] > (datetime(_year, _month, _day, 18, 5) - timedelta(1)):
            _ += 1
            __ += item[1]
        
        day_mean = round(__ / _, 3)

        # day_mean = round(reduce_interval / count_interval, 3)
        
        check_point[7] = float(row["close"])
        check_point[8] = day_mean

        delta = round(float(row["close"]) - day_mean, 3)

        check_point[9] = delta
        check_point[10] = single_point
        check_point[11] = delta * single_point
        
        reduce_day.append(delta)
        graph_point[6] = delta * single_point

        count_interval = 0
        reduce_interval = 0

        graph_point[0] = graph_point[0][:-9]
        graph_points.append(graph_point[:])
        delta_list.append(graph_point[:])

        stack = []

    check_points.append(check_point)

  # for duration in durations:
  #   _sub_list = reduce_day[-duration:]
  #   try:
  #     insertValue(out_dict, round(sum(_sub_list) / len(_sub_list), 3))
  #   except:
  #     insertValue(out_dict, "N/A")
  # try:
  #   insertValue(out_dict, round(sum(reduce_day) / len(reduce_day), 3))
  # except:
  #   insertValue(out_dict, "N/A")

  try:
    df = pd.DataFrame(np.array(delta_list)[:, 6].tolist(), columns=["delta"])
    df_wma = df["delta"].rolling(252).apply(lambda prices: np.round(np.dot(prices, weights)/weights.sum(), 3), raw=True)
    wma = df_wma.tolist()

    for index in range(len(wma)):
      graph_points[-(index + 1)][7] = wma[-(index + 1)]
  except:
    pass

  summary_stat.append([symbol, wma[-1], np.round(df_wma.mean(skipna=True), 3)])

  return days

def writeResult(_dict):
  ## wirte the result as xlsx file using openpyxl.

  print("writing...")
  def _get_column_num(number):

    new_number = number
    new_number += 4 if number > 0 else 0
    # new_number += 1 if number > 4 else 0
    # new_number += 1 if number > 7 else 0
    return new_number

  book = openpyxl.Workbook()
  sheet = book.active

  symbol_count = len(_dict.keys())

  ## write the first row.
  for index, key in enumerate(returnDict().keys()):
    sheet.cell(row=2, column=_get_column_num(index + 1 + 1)).value = key

  ## write each values.
  for symbol_index, symbol_key in enumerate(_dict):
    for column_index, column_key in enumerate(_dict[symbol_key]):
      basic_row = symbol_index + 2 + 1
      basic_col = _get_column_num(column_index + 1 + 1)

      for index, item in enumerate(_dict[symbol_key][column_key]):
        real_row = basic_row + index * (symbol_count + 2)
        sheet.cell(row=real_row, column=basic_col).value = item

        if column_index == 0:
          sheet.cell(row=real_row, column=_get_column_num(1)).value = symbol_key

          if symbol_index == 0:
            sheet.cell(row=real_row - 1, column=_get_column_num(1)).value = months[index]

  ## write to file.
  result_path = out_path + "/result.xlsx"

  if os.path.exists(result_path):
    os.remove(result_path)

  os.makedirs(out_path, exist_ok=True)
  
  book.save(result_path)
      

if __name__ == "__main__":
  ## path for the input file.
  root_path = "./input"
  out_path = "./output"
  path = "/AutoData/iqfeed_year_data_20190216-20210216.csv"

  ## prepare the source raw data.
  df_data = pd.read_csv(root_path + path)
  df_data = df_data[["timestamp", "high", "low", "open", "close", "symbol"]]

  ## prepare BuyStrength_SellWeakness data.
  df_additional = read_ods(root_path + "/BuyStrength_SellWeakness.ods", 1)
  df_additional = df_additional.iloc[2:, :]
  df_additional = df_additional[df_additional["unnamed.2"].notnull()].reset_index(drop=True)

  ## prepare IQFeed_Symbols data.
  df_singlepoint = pd.read_excel(root_path + "/IQFeed_Symbols.xlsx", engine="openpyxl")
  df_singlepoint = df_singlepoint[df_singlepoint["Symbol"].notnull()].reset_index(drop=True)

  ## prepare variables to save data.
  graph_points = []
  precious_dict = {}
  item_lens = []
  check_points = []

  summary_stat = []

  weights = []
  for i in range(252):
    _ = 1 / 252
    _ += 1 / 63 if i >= 252 - 63 else 0
    _ += 1 / 21 if i >= 252 - 20 else 0
    _ += 1 / 5 if i >= 252 - 5 else 0

    weights.append(_)
  weights = np.array(weights)
  print(weights)

  ## set start date.
  start_date = datetime(2019, 2, 16)

  for index in tqdm(range(min(len(df_additional.index), 2000))):

    row = df_additional.iloc[index, :]
    _time_str = row.iloc[4]
    _product_name = row.iloc[1]
    _row_single = df_singlepoint[df_singlepoint["Name"] == _product_name]
    _product_code = row.iloc[0]

    _hour = int(_time_str[2:_time_str.find('H')])
    _min = int(_time_str[_time_str.find('H') + 1: _time_str.find('M')])

    settlement = str(time(_hour, _min))
    pre_settlement = str(time(23 if _hour == 0 else _hour - 1, _min))
    open_time = time(18, 5)

    symbol = _row_single["Symbol"].values[0]
    single_point = _row_single["Single point value"].values[0]
    threshold = row.iloc[5]
    threshold = threshold * 100

    precious_dict.update({_product_code: returnDict()})
    data = df_data[df_data["symbol"] == symbol].reset_index(drop=True)

    item_lens.append([_product_code, calculateValues(weights, check_points, graph_points, precious_dict[_product_code], data, single_point, settlement, open_time, threshold, symbol, summary_stat)])

  # writeResult(precious_dict)
  ## output checkpoint file.
  df = pd.DataFrame(check_points, columns=["timestamp", "high", "low", "open", "close", "symbol", "mean", "close", "day_mean", "delta", "single_point", "delta in dollars"])
  df.to_csv(out_path + "/check_points.csv", index=False)

  ## output graph point file.
  df = pd.DataFrame(graph_points, columns=["timestamp", "high", "low", "open", "close", "symbol", "delta", "TWAP"])
  df_summary = pd.DataFrame(summary_stat, columns=["symbol", "TWAP", "average TWAP"])

  
  df_summary.to_excel(out_path + "/summary stat.xlsx")
  df.to_excel(out_path + "/graph_points.xlsx", index=False)

  ## draw graph on graph point file.
  book = openpyxl.load_workbook(out_path + "/graph_points.xlsx")
  sheet = book.active
  graph_sheet = book.create_sheet("graph")

  row_over = 0
  for index, item in enumerate(item_lens):

    chart = LineChart()
    chart.title = item[0]
    chart.x_axis.title = "Time"
    chart.y_axis.title = "TWAP"
    chart.style = 2
    x = Reference(sheet, min_col=1, min_row=row_over+2, max_row=row_over+item[1]+1)
    # y = Reference(sheet, min_col=8, min_row=row_over+2, max_col=8, max_row=row_over+item[1]+1)
    values = Reference(sheet, min_col=8, min_row=row_over+2, max_col=8, max_row=row_over+item[1]+1)
    # values = Series(y, xvalues=x)
    chart.add_data(values)
    chart.set_categories(x)
    chart.height = 20
    chart.width = 40

    graph_sheet.add_chart(chart, "A" + str(index * 50 + 1))
    row_over += item[1]

  book.save(out_path + "/graph_points.xlsx")

  print("done!")